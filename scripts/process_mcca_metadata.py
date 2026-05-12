#!/usr/bin/env python3
"""
process_mcca_metadata.py
========================

Reads MCCA-CellLineAnnotations-2025Q3.xlsx and emits
web_data/metadata.json with the per-cell-line annotation we need for
MouseCLB. Also emits curated-list overlap and gap reports so the build
stays honest about which lines are or aren't covered.

Output schema (web_data/metadata.json):
  {
    "cellLines": ["MCCA0026", ...],          // canonical MCCA-ID for now
    "names": {"MCCA0026": "4T1", ...},
    "lineage": {"MCCA0026": "MammaryGland", ...},
    "cancerType": {"MCCA0026": "MammaryCancer", ...},
    "site": {"MCCA0026": "Primary", ...},
    "modelType": {"MCCA0026": "Wildtype", ...},
    "mouseModel": {"MCCA0026": "BALB/cJ", ...},
    "strain": {"MCCA0026": "BALB/cJ", ...},
    "gender": {"MCCA0026": "Female", ...},
    "mhcA": {"MCCA0026": "H2-Kd", ...},
    "mhcB": {"MCCA0026": "H2-Db", ...},
    "media": {"MCCA0026": "RPMI...", ...},
    "morphology": {"MCCA0026": "Epithelial", ...},
    "survivalDays": {"MCCA0026": 21, ...},
    "metastasis": {"MCCA0026": "Yes", ...},
    "wgd": {"MCCA0026": false, ...},          // ComplexRearrangement proxy
    "chromothripsis": {"MCCA0026": false, ...},
    "immunocompetentTransplantation": {"MCCA0026": true, ...},
    "pmid": {"MCCA0026": "12345678", ...},
    "tumorLocation": {"MCCA0026": "MammaryGland", ...}
  }

The MCCA file lists 999 rows / 588 unique CellLineNames. We keep the MCCA-ID as
the canonical identifier (clones / derivatives get distinct IDs). The app's
default view will hide IntestinalHealthyTissue and similar non-cancer entries.
"""

import json
import os
from collections import defaultdict
import openpyxl

# Curated target list — see CURATED_CELL_LINES.md. Match CellLineName
# substrings case-insensitively. Each canonical name maps to a tier.
CURATED = {
    # Tier 1
    'B16-F10':   {'tier': 1, 'aliases': ['B16-F10', 'B16F10', 'B16-F10M']},
    'MC38':      {'tier': 1, 'aliases': ['MC38', 'MC-38']},
    'CT26':      {'tier': 1, 'aliases': ['CT26', 'CT-26']},
    '4T1':       {'tier': 1, 'aliases': ['4T1', '4T1-T0']},
    'LL/2':      {'tier': 1, 'aliases': ['LL/2', 'LL2', 'LLC', 'LLC1']},
    'EMT6':      {'tier': 1, 'aliases': ['EMT6', 'EMT-6']},
    'Renca':     {'tier': 1, 'aliases': ['RENCA', 'Renca']},
    'Pan02':     {'tier': 1, 'aliases': ['Pan02', 'PAN02', 'PANC02']},
    'EL4':       {'tier': 1, 'aliases': ['EL4', 'EL-4']},
    'Hepa1-6':   {'tier': 1, 'aliases': ['HEPA1-6', 'Hepa1-6', 'Hepa16']},
    'E0771':     {'tier': 1, 'aliases': ['E0771', 'EO771']},
    'MOC1':      {'tier': 1, 'aliases': ['MOC1']},
    'MOC2':      {'tier': 1, 'aliases': ['MOC2']},
    'A20':       {'tier': 1, 'aliases': ['A20']},
    'YAC-1':     {'tier': 1, 'aliases': ['YAC1', 'YAC-1']},
    'ID8':       {'tier': 1, 'aliases': ['ID8']},
    'TC-1':      {'tier': 1, 'aliases': ['TC-1', 'TC1']},
    'B16-F1':    {'tier': 1, 'aliases': ['B16-F1', 'B16F1']},
    '67NR':      {'tier': 1, 'aliases': ['67NR']},
    # Tier 2 (still surface from MCCA when present so user can find them)
    '4T07':      {'tier': 2, 'aliases': ['4T07']},
    '168FARN':   {'tier': 2, 'aliases': ['168FARN']},
    'AT-3':      {'tier': 2, 'aliases': ['AT-3', 'AT3']},
    'GL261':     {'tier': 2, 'aliases': ['GL261', 'GL-261']},
    'MB49':      {'tier': 2, 'aliases': ['MB49']},
    'RM-1':      {'tier': 2, 'aliases': ['RM-1', 'RM1']},
    'RM-9':      {'tier': 2, 'aliases': ['RM-9', 'RM9']},
    'AB1':       {'tier': 2, 'aliases': ['AB1']},
}

# Columns we want to keep
KEEP_COLS = [
    'MCCA-ID', 'CellLineName', 'PMID', 'PublicationStatus',
    'MouseID', 'TumorLocation', 'MouseModelType', 'MouseModel', 'MouseModelDetailed',
    'Tissue', 'Lineage', 'Site', 'CancerType', 'CancerTypeDetailed',
    'Media', 'CellCultureSystem',
    'MicroscopicMorphology', 'MicroscopicMorphologyDetailed',
    'SurvivalDays', 'DistantMetastasis',
    'ComplexRearrangement', 'Chromothripsis',
    'StrainRank1Name', 'StrainRank1Percentage',
    'StrainRank2Name', 'StrainRank2Percentage',
    'MhcHaplotypeA', 'MhcHaplotypeB',
    'Gender', 'ImmunocompetentTransplantation',
    'CellLineSource', 'CellLineDistributor',
]

def find_curated_match(cell_line_name):
    """Return canonical curated name if the line matches, else None."""
    if not cell_line_name:
        return None
    cn_up = str(cell_line_name).strip().upper()
    for canonical, info in CURATED.items():
        for alias in info['aliases']:
            if cn_up == alias.upper():
                return canonical
    return None

def _val(x):
    """Normalise cell value: empty strings, NaN-like and Excel sentinels → None.
    Numbers stay as numbers; everything stringy is stripped. The names column
    sometimes contains numeric-looking entries (e.g. 67NR could be loaded as
    a float by openpyxl if the cell has number format) — we coerce those
    back to strings so downstream code can rely on str() semantics."""
    if x is None:
        return None
    if isinstance(x, str):
        s = x.strip()
        if not s or s.upper() in ('NA', 'N/A', 'NAN', '-'):
            return None
        return s
    return x

def _str(x):
    """Force-to-string version of _val for fields we render as text (names,
    lineages, etc.) — coerces stray numeric cells to clean strings."""
    v = _val(x)
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v) if not isinstance(v, str) else v

def main():
    here = os.path.dirname(os.path.abspath(__file__))
    src = os.path.join(here, '..', '..', 'correlate_v2', 'mouse_cell_lines',
                       'MCCA-CellLineAnnotations-2025Q3.xlsx')
    out_dir = os.path.join(here, '..', 'web_data')
    os.makedirs(out_dir, exist_ok=True)

    print(f'Reading {src}')
    wb = openpyxl.load_workbook(src, read_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    header = list(next(it))
    rows = [r for r in it]
    print(f'  {len(rows)} rows × {len(header)} columns')

    col_idx = {c: i for i, c in enumerate(header)}

    # Build the metadata structure keyed by MCCA-ID
    meta = {
        'cellLines': [],
        'names': {}, 'pmid': {}, 'tumorLocation': {}, 'modelType': {},
        'mouseModel': {}, 'mouseModelDetailed': {},
        'tissue': {}, 'lineage': {}, 'site': {}, 'cancerType': {}, 'cancerTypeDetailed': {},
        'media': {}, 'cultureSystem': {},
        'morphology': {}, 'morphologyDetailed': {},
        'survivalDays': {}, 'metastasis': {},
        'complexRearrangement': {}, 'chromothripsis': {},
        'strain': {}, 'strainPct': {},
        'mhcA': {}, 'mhcB': {},
        'gender': {}, 'immunocompetent': {},
        'source': {}, 'distributor': {},
        'curated': {},     # MCCA-ID → canonical curated name (only if matched)
        'curatedTier': {}, # MCCA-ID → tier (1 or 2) if curated
    }

    counts = defaultdict(int)
    curated_hits = defaultdict(list)
    for r in rows:
        cl = _str(r[col_idx['MCCA-ID']])
        name = _str(r[col_idx['CellLineName']])
        if not cl:
            continue
        meta['cellLines'].append(cl)
        if name is not None: meta['names'][cl] = name
        v = _val(r[col_idx['PMID']]);                meta['pmid'][cl]                  = v if v else None
        v = _val(r[col_idx['TumorLocation']]);       meta['tumorLocation'][cl]         = v
        v = _val(r[col_idx['MouseModelType']]);      meta['modelType'][cl]             = v
        v = _val(r[col_idx['MouseModel']]);          meta['mouseModel'][cl]            = v
        v = _val(r[col_idx['MouseModelDetailed']]);  meta['mouseModelDetailed'][cl]    = v
        v = _val(r[col_idx['Tissue']]);              meta['tissue'][cl]                = v
        v = _val(r[col_idx['Lineage']]);             meta['lineage'][cl]               = v
        v = _val(r[col_idx['Site']]);                meta['site'][cl]                  = v
        v = _val(r[col_idx['CancerType']]);          meta['cancerType'][cl]            = v
        v = _val(r[col_idx['CancerTypeDetailed']]);  meta['cancerTypeDetailed'][cl]    = v
        v = _val(r[col_idx['Media']]);               meta['media'][cl]                 = v
        v = _val(r[col_idx['CellCultureSystem']]);   meta['cultureSystem'][cl]         = v
        v = _val(r[col_idx['MicroscopicMorphology']]);         meta['morphology'][cl]         = v
        v = _val(r[col_idx['MicroscopicMorphologyDetailed']]); meta['morphologyDetailed'][cl] = v
        v = _val(r[col_idx['SurvivalDays']]);        meta['survivalDays'][cl]          = v
        v = _val(r[col_idx['DistantMetastasis']]);   meta['metastasis'][cl]            = v
        v = _val(r[col_idx['ComplexRearrangement']]); meta['complexRearrangement'][cl] = v
        v = _val(r[col_idx['Chromothripsis']]);      meta['chromothripsis'][cl]        = v
        v = _val(r[col_idx['StrainRank1Name']]);     meta['strain'][cl]                = v
        v = _val(r[col_idx['StrainRank1Percentage']]); meta['strainPct'][cl]           = v
        v = _val(r[col_idx['MhcHaplotypeA']]);       meta['mhcA'][cl]                  = v
        v = _val(r[col_idx['MhcHaplotypeB']]);       meta['mhcB'][cl]                  = v
        v = _val(r[col_idx['Gender']]);              meta['gender'][cl]                = v
        v = _val(r[col_idx['ImmunocompetentTransplantation']]); meta['immunocompetent'][cl] = v
        v = _val(r[col_idx['CellLineSource']]);      meta['source'][cl]                = v
        v = _val(r[col_idx['CellLineDistributor']]); meta['distributor'][cl]           = v

        counts[meta['lineage'].get(cl) or 'unknown'] += 1
        # Curated overlap
        canon = find_curated_match(name)
        if canon:
            meta['curated'][cl] = canon
            meta['curatedTier'][cl] = CURATED[canon]['tier']
            curated_hits[canon].append((cl, name))

    # Sort cellLines so the curated lines come first (so the default app list
    # leads with the workhorses), then alphabetical by name within each bucket.
    def _sort_key(cl):
        tier = meta['curatedTier'].get(cl, 99)
        return (tier, (meta['names'].get(cl) or '').upper())
    meta['cellLines'].sort(key=_sort_key)

    print(f"\n  Wrote {len(meta['cellLines'])} cell lines.")
    print(f"  Lineage distribution (top 10):")
    for k, v in sorted(counts.items(), key=lambda x: -x[1])[:10]:
        print(f"    {v:4d}  {k}")

    print(f"\n  Curated-list MCCA coverage:")
    for canon, info in CURATED.items():
        hits = curated_hits.get(canon, [])
        marker = '✓' if hits else '✗'
        names_str = ', '.join(f'{cl}={name}' for cl, name in hits) or 'NOT IN MCCA'
        print(f"    {marker}  T{info['tier']}  {canon:12s}  →  {names_str}")

    out_path = os.path.join(out_dir, 'metadata.json')
    with open(out_path, 'w') as f:
        json.dump(meta, f, separators=(',', ':'))
    print(f"\n  → {out_path}  ({os.path.getsize(out_path):,} bytes)")

if __name__ == '__main__':
    main()
