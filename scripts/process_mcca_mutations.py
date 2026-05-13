#!/usr/bin/env python3
"""Process MCCA's mutation file into per-cell-line driver-mutation summaries.

Input:  MCCA-Mutations-2025Q3.xlsx (downloaded from https://www.mcca.tum.de;
        one row per variant × annotation).

Output: web_data/mutations.json with two layers per cell line:
  - aggregate counts (HIGH-impact, MODERATE-impact)
  - per-driver-gene variant detail (gene, impact, effect, HGVS_p, position)

The driver-gene panel is a ~60-gene cancer-driver list (mouse symbols
matching the most-studied tumour suppressors and oncogenes); non-driver
HIGH-impact mutations are summarised only as a count, so the JSON stays
small (~1-2 MB instead of the raw 10 MB Excel).

Why not include MODIFIER / LOW rows: MODIFIER is non-coding / pseudogene
(108k rows, 82% of the file) and rarely actionable in a cell-line context;
LOW is mostly synonymous. HIGH and MODERATE together cover the biologically
meaningful changes (~16k rows).

Re-run when MCCA publishes a new mutation release.
"""
import json, os, sys
from collections import Counter, defaultdict

try:
    import openpyxl
except ImportError:
    print('install: pip install openpyxl', file=sys.stderr); sys.exit(1)

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
IN_PATH = os.environ.get('MCCA_MUTATIONS_XLSX',
    '/Users/fredrikwermeling/Documents/correlate_v2/mouse_cell_lines/MCCA-Mutations-2025Q3.xlsx')
OUT_PATH = os.path.join(ROOT, 'web_data', 'mutations.json')

# Curated cancer-driver gene panel (mouse symbols). Pulled from COSMIC
# Cancer Gene Census tier-1 with mouse-orthologue mapping, plus the
# well-known TGF-β / NOTCH / DDR pathway members.
DRIVER_GENES = {
    # Tumour suppressors
    'Trp53', 'Apc', 'Pten', 'Rb1', 'Brca1', 'Brca2', 'Cdkn2a', 'Cdkn2b',
    'Vhl', 'Nf1', 'Nf2', 'Tsc1', 'Tsc2', 'Smad4', 'Smad2', 'Stk11',
    'Setd2', 'Atrx', 'Pbrm1', 'Kmt2c', 'Kmt2d', 'Asxl1', 'Bap1', 'Bcor',
    'Crebbp', 'Ep300', 'Mlh1', 'Msh2', 'Msh6', 'Pms2', 'Notch2', 'Fbxw7',
    'Wt1', 'Foxa1', 'Cdh1', 'Spop', 'Stag2',
    # Oncogenes
    'Kras', 'Hras', 'Nras', 'Braf', 'Myc', 'Mycn', 'Mycl1', 'Egfr',
    'Erbb2', 'Erbb3', 'Met', 'Alk', 'Ret', 'Kit', 'Pdgfra', 'Flt3',
    'Jak1', 'Jak2', 'Akt1', 'Pik3ca', 'Pik3cb', 'Mtor', 'Notch1',
    'Ctnnb1', 'Idh1', 'Idh2', 'Bcl2', 'Bcl6', 'Mdm2', 'Mdm4', 'Ezh2',
    'Sox2',
    # DNA-damage / replication
    'Pole', 'Pold1', 'Atm', 'Atr', 'Chek1', 'Chek2', 'Mre11a', 'Rad51',
    'Nbn'
}

def main():
    print(f'reading {IN_PATH} ...')
    wb = openpyxl.load_workbook(IN_PATH, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    c = {name: i for i, name in enumerate(header)}

    by_cl = defaultdict(lambda: {
        'totalHigh': 0,
        'totalModerate': 0,
        'driverMuts': [],
        # Track gene-level non-driver HIGH-impact totals so we can flag
        # interesting outliers later if we want (kept compact).
        'highImpactGenes': Counter()
    })
    seen_pos = defaultdict(set)  # dedupe by (cl, chrom, pos, gene)

    n = 0
    for row in rows:
        n += 1
        impact = row[c['ANN[*].IMPACT']]
        if impact not in ('HIGH', 'MODERATE'):
            continue
        gene = row[c['ANN[*].GENE']]
        gene = str(gene) if gene is not None else ''
        if not gene or gene.startswith('Gm') or gene.endswith('Rik'):
            continue
        cl = row[c['MCCA_ID']]
        chrom = row[c['CHROM']]
        pos = row[c['POS-mm10']]
        dedup_key = (chrom, pos, gene)
        if dedup_key in seen_pos[cl]:
            continue
        seen_pos[cl].add(dedup_key)

        e = by_cl[cl]
        if impact == 'HIGH':
            e['totalHigh'] += 1
            e['highImpactGenes'][gene] += 1
        elif impact == 'MODERATE':
            e['totalModerate'] += 1

        if gene in DRIVER_GENES:
            effect = row[c['ANN[*].EFFECT']]
            hgvs_p = row[c['ANN[*].HGVS_P']]
            af = row[c['GEN[Tumor].AF']]
            e['driverMuts'].append({
                'gene': gene,
                'impact': impact,
                'effect': effect,
                'hgvsP': hgvs_p,
                'chrom': str(chrom) if chrom is not None else None,
                'pos': int(pos) if isinstance(pos, (int, float)) else None,
                'af': round(float(af), 3) if isinstance(af, (int, float)) else None
            })

    # Convert Counter → top-5 most-mutated non-driver genes per cell line.
    for cl, e in by_cl.items():
        top = e['highImpactGenes'].most_common(5)
        e['topHighImpactGenes'] = [{'gene': g, 'n': n} for g, n in top]
        del e['highImpactGenes']
        # Sort drivers by impact then gene
        e['driverMuts'].sort(key=lambda m: (m['impact'] != 'HIGH', m['gene']))

    out = {
        '_doc': 'Per-cell-line driver-mutation summary derived from MCCA-Mutations-2025Q3.xlsx (https://www.mcca.tum.de). Filtered to HIGH and MODERATE SnpEff impact; predicted-non-coding (MODIFIER) and synonymous (LOW) rows dropped. Per-line: totalHigh / totalModerate counts (rough TMB proxy), full variant detail for the ~70-gene cancer-driver panel, and the top-5 most-mutated non-driver genes by HIGH-impact count.',
        'source': {
            'name': 'MCCA',
            'file': 'MCCA-Mutations-2025Q3.xlsx',
            'url':  'https://www.mcca.tum.de'
        },
        'driverPanel': sorted(DRIVER_GENES),
        'byCellLine': dict(by_cl)
    }
    with open(OUT_PATH, 'w') as f:
        json.dump(out, f, indent=1)
    print(f'wrote {OUT_PATH}')
    print(f'  rows scanned: {n}')
    print(f'  cell lines with driver-panel hits: {sum(1 for e in by_cl.values() if e["driverMuts"])}')
    print(f'  cell lines with any HIGH-impact: {sum(1 for e in by_cl.values() if e["totalHigh"])}')

if __name__ == '__main__':
    main()
